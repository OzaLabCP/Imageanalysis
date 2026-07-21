"""Interactive Excel workbook from a CellScope measurements parquet.

Unlike the PNG figures (flat, standalone artifacts), this writes an ``.xlsx`` with
**live formulas** and **native Excel charts**: the per-cell values live on a data
sheet, every summary number is a ``COUNTIFS``/``AVERAGEIFS`` formula, and the
responder threshold is a single editable cell - change it and every ``% responder``
figure and every chart recomputes. The charts are real Excel chart objects, so
they can be restyled, hovered, and pasted into a paper, not exported images.

Sheets
------
* **Settings**    - editable responder threshold + endpoint timepoint (yellow).
* **Cells**       - the raw per-cell rows the formulas read (condition, well,
                    timepoint, intensity).
* **ResponderPct**- %% responders per condition x timepoint (live formulas).
* **MeanGreen**   - mean intensity per condition x timepoint (live formulas).
* **PerWell**     - %% responders per well at the endpoint (the honest unit).
* **Charts**      - native line + bar charts referencing the sheets above.

Needs ``openpyxl`` (the ``analysis`` extra).
"""

from __future__ import annotations

_MAX_CELL_ROWS = 400_000  # keep the workbook openable; note if we sample below this


def _col(idx: int) -> str:
    """1-based column index -> Excel letter(s)."""
    s = ""
    while idx > 0:
        idx, r = divmod(idx - 1, 26)
        s = chr(65 + r) + s
    return s


def build_workbook(parquet, out_xlsx, channel=None, platemap=None, threshold=None,
                   min_diameter=6.0, max_eccentricity=0.95):
    """Write an interactive Excel report. Returns a small summary dict."""
    import numpy as np
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.styles import Alignment, Font, PatternFill

    from cellscope.analyze import GREEN_DEFAULT, _gate, _load, _responder_threshold

    raw, channel = _load(parquet, platemap, channel or GREEN_DEFAULT)
    df = _gate(raw, channel, min_diameter, max_eccentricity)
    if "Well" not in df.columns or "Timepoint" not in df.columns:
        raise ValueError("parquet needs Well and Timepoint columns")
    if threshold is None:
        threshold = float(_responder_threshold(df[channel]))

    df = df[np.isfinite(df[channel])].copy()
    df["group"] = df["group"].astype(str)
    df["Well"] = df["Well"].astype(str)
    groups = sorted(df["group"].unique())
    tvals = sorted(int(t) for t in df["Timepoint"].unique())
    endpoint = tvals[-1]
    wells = (df[["Well", "group"]].drop_duplicates()
             .sort_values(["group", "Well"]).values.tolist())

    sampled = False
    if len(df) > _MAX_CELL_ROWS:
        df = df.sample(_MAX_CELL_ROWS, random_state=0)
        sampled = True

    wb = Workbook()
    # Arial throughout via the default (Normal) style - cheap even for 100k rows.
    try:
        wb._named_styles["Normal"].font = Font(name="Arial", size=10)
    except Exception:  # pragma: no cover - style internals differ across versions
        pass
    hdr = Font(name="Arial", size=10, bold=True)
    title = Font(name="Arial", size=12, bold=True)
    edit_fill = PatternFill("solid", fgColor="FFF3B0")
    pct_fmt = "0.0%"

    # --- Settings ---------------------------------------------------------
    st = wb.active
    st.title = "Settings"
    st["A1"] = "CellScope interactive report"; st["A1"].font = title
    rows = [
        ("Setting", "Value"),
        ("Responder threshold (intensity)", round(threshold, 2)),
        ("Channel", channel),
        ("Endpoint timepoint", endpoint),
    ]
    for i, (a, b) in enumerate(rows, start=3):
        st[f"A{i}"] = a; st[f"B{i}"] = b
    st["A3"].font = st["B3"].font = hdr
    for c in ("B4", "B6"):  # threshold + endpoint are the editable levers
        st[c].fill = edit_fill
    st["A8"] = ("Yellow cells are editable. Change the threshold or endpoint and "
                "every % responder figure and chart below updates automatically.")
    st["A8"].alignment = Alignment(wrap_text=True)
    st.column_dimensions["A"].width = 34
    st.column_dimensions["B"].width = 22
    THR = "Settings!$B$4"
    END = "Settings!$B$6"

    # --- Cells (raw data the formulas read) -------------------------------
    ce = wb.create_sheet("Cells")
    ce.append(["Condition", "Well", "Timepoint", channel])
    for c in ce[1]:
        c.font = hdr
    for cond, well, tp, val in df[["group", "Well", "Timepoint", channel]].itertuples(
            index=False, name=None):
        ce.append([cond, well, int(tp), float(val)])
    ce.freeze_panes = "A2"
    # Cells columns: A=Condition B=Well C=Timepoint D=intensity. Bound every
    # COUNTIFS/AVERAGEIFS to the actual data rows (not whole columns) - whole-column
    # references scan ~1M rows per call and make recalculation crawl.
    nrow = ce.max_row
    A = f"Cells!$A$2:$A${nrow}"
    B = f"Cells!$B$2:$B${nrow}"
    C = f"Cells!$C$2:$C${nrow}"
    D = f"Cells!$D$2:$D${nrow}"

    def _matrix(sheet_name, title_text, kind):
        """A timepoint (rows) x condition (cols) matrix of live formulas."""
        ws = wb.create_sheet(sheet_name)
        ws["A1"] = title_text; ws["A1"].font = title
        ws["A2"] = "Timepoint"; ws["A2"].font = hdr
        for j, g in enumerate(groups):
            cell = ws.cell(row=2, column=2 + j, value=g); cell.font = hdr
        for i, t in enumerate(tvals):
            ws.cell(row=3 + i, column=1, value=int(t))
            for j in range(len(groups)):
                gcol = _col(2 + j)
                tref = f"$A{3 + i}"
                gref = f"{gcol}$2"
                if kind == "pct":
                    f = (f'=IFERROR(COUNTIFS({A},{gref},{C},{tref},{D},">"&{THR})'
                         f'/COUNTIFS({A},{gref},{C},{tref}),"")')
                    fmt = pct_fmt
                else:
                    f = f'=IFERROR(AVERAGEIFS({D},{A},{gref},{C},{tref}),"")'
                    fmt = "0"
                cell = ws.cell(row=3 + i, column=2 + j, value=f)
                cell.number_format = fmt
        ws.column_dimensions["A"].width = 12
        return ws, 2, 2 + len(groups) - 1, 3, 2 + len(tvals)

    rp, c0, c1, r0, r1 = _matrix("ResponderPct",
                                 "% responders by condition over time (live)", "pct")
    mg, *_ = _matrix("MeanGreen", "Mean intensity by condition over time (live)", "mean")

    # --- PerWell (the unit that is honestly n) ----------------------------
    pw = wb.create_sheet("PerWell")
    pw["A1"] = "% responders per well at endpoint (each row = one well)"
    pw["A1"].font = title
    pw["A2"] = "Well"; pw["B2"] = "Condition"; pw["C2"] = "% responders (endpoint)"
    for c in (pw["A2"], pw["B2"], pw["C2"]):
        c.font = hdr
    for i, (well, cond) in enumerate(wells):
        r = 3 + i
        pw.cell(row=r, column=1, value=well)
        pw.cell(row=r, column=2, value=cond)
        f = (f'=IFERROR(COUNTIFS({A},$B{r},{B},$A{r},{C},{END},{D},">"&{THR})'
             f'/COUNTIFS({A},$B{r},{B},$A{r},{C},{END}),"")')
        c = pw.cell(row=r, column=3, value=f); c.number_format = pct_fmt
    pw.column_dimensions["A"].width = 12
    pw.column_dimensions["B"].width = 18
    pw.column_dimensions["C"].width = 22
    pw_last = 2 + len(wells)

    # --- Charts (native, referencing the live sheets) ---------------------
    ch = wb.create_sheet("Charts")

    def _line(anchor, src, ttl, y):
        c = LineChart(); c.title = ttl; c.style = 2
        c.x_axis.title = "Timepoint"; c.y_axis.title = y
        c.height, c.width = 8.5, 16
        data = Reference(src, min_col=c0, max_col=c1, min_row=r0 - 1, max_row=r1)
        cats = Reference(src, min_col=1, min_row=r0, max_row=r1)
        c.add_data(data, titles_from_data=True)
        c.set_categories(cats)
        ch.add_chart(c, anchor)

    _line("A1", rp, "% responders over time", "% responders")
    _line("A19", mg, "Mean intensity over time", f"mean {channel}")

    bar = BarChart(); bar.title = "% responders per well (endpoint)"; bar.style = 10
    bar.y_axis.title = "% responders"; bar.x_axis.title = "Well"
    bar.height, bar.width = 8.5, 16
    bdata = Reference(pw, min_col=3, min_row=2, max_row=pw_last)
    bcats = Reference(pw, min_col=1, min_row=3, max_row=pw_last)
    bar.add_data(bdata, titles_from_data=True)
    bar.set_categories(bcats)
    bar.legend = None
    ch.add_chart(bar, "A37")

    wb.save(out_xlsx)
    return {"path": out_xlsx, "channel": channel, "threshold": float(threshold),
            "conditions": groups, "timepoints": tvals, "wells": len(wells),
            "cells_written": int(len(df)), "sampled": sampled}
